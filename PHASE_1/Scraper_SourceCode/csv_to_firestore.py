"""
Script that converts scraped data in a .csv format to the Thicc Peas' DB.

This script will send the diseases and all reports and articles.
"""

from openpyxl import load_workbook
from firestore_client import FireStore_Client
from datetime import datetime
import os

class scraped_data_xlsx_to_firestore:
    """
    Sends data from the 'scraped_data' xlsx file to the FireStore DB.
    """
    def __init__(self, path_to_xlsx: str, debug = True):
        self.client  = FireStore_Client()

        if not os.path.exists(path_to_xlsx):
            raise FileNotFoundError(f"'{path_to_xlsx}' not found! Please add it.")
        self.workbook = load_workbook(filename=path_to_xlsx, data_only=True)

        # Makes this class print information about what it's sending.
        self.debug = debug

    def send_diseases(self, disease_name = None):
        """
        Sends all diseases from the 'Diseases' sheet to the FireStore.
        """
        # Set active sheet to 'Diseases'
        disease_sheet = self.workbook["Diseases"]

        # Loop each row starting from row '2' till the first non-null row.
        print("==========Sending 'Diseases' Sheet==========")
        for row in disease_sheet.iter_rows(2, 0):
            disease_id = row[0].value
            name = row[1].value
            symptoms = row[2].value

            if disease_name is not None and name != disease_name:
                continue;

            # Parse symptoms (which should be of form "<symptom_1>|<symptom_2>|...") into an array
            parsed_symptoms = symptoms.split("|")

            if disease_id is not None:
                print(disease_id)
                print(name)
                # If the row is not empty then send the disease to the FireStore.
                self.client.write_disease(disease_id, name, parsed_symptoms)
        print("==========        COMPLETE        ==========")

    def send_reports_and_articles(self, disease_name = None):
        """
        Sends all reports and articles to the FireStore for each disease in the xlsx file.
        """
        disease_report_and_article_sheets = self.workbook.sheetnames
        disease_report_and_article_sheets.remove("Diseases")

        print("==========Sending 'Reports' and 'Articles'==========")
        for sheet_name in disease_report_and_article_sheets:
            if disease_name is not None and sheet_name != disease_name:
                continue

            print(f"======{sheet_name}=====")
            for row in self.workbook[sheet_name].iter_rows(2):

                # Send Article
                placeholder_article_id          = row[0].value
                url                 = row[1].value
                headline            = row[2].value
                main_text           = row[3].value
                date_of_publication = row[4].value

                parsed_date_of_publication = date_of_publication
                if (date_of_publication is not None and not isinstance(date_of_publication, datetime)):
                    parsed_date_of_publication = datetime.strptime(date_of_publication, "%Y-%m-%d")

                article_id_hash = None
                if placeholder_article_id is not None:
                    # If the row is not empty then send the article to the FireStore.
                    article_id_hash = self.client.write_article_auto_id(url, headline, main_text, parsed_date_of_publication)

                # Send Report
                placeholder_report_id        = row[6].value
                # article_id       = row[7].value
                disease_id       = row[8].value
                event_date       = row[9].value
                location         = row[10].value
                reported_cases   = row[12].value
                hospitalisations = row[13].value
                deaths           = row[14].value

                parsed_event_date = event_date
                if (event_date is not None and not isinstance(event_date, datetime)):
                    parsed_event_date = datetime.strptime(event_date, "%Y-%m-%d")

                if placeholder_report_id is not None:
                    # If the row is not empty then send the report to the FireStore.
                    self.client.write_report_auto_id(article_id_hash, disease_id, location, parsed_event_date, reported_cases, hospitalisations, deaths)

if __name__ == "__main__":
    # Initialise sender
    try:
        sender = scraped_data_xlsx_to_firestore("./Scraping_Sheet.xlsx")
    except FileNotFoundError as e:
        print("Could not initialise the sender.")
        print(e)
        exit(1)

    # Send diseases
    sender.send_diseases()

    # Send reports and articles for all diseases
    sender.send_reports_and_articles()

    # Send only cholera and chikungunya disease and reports.
    # sender.send_diseases("cholera")
    # # sender.send_diseases("chikungunya")
    # sender.send_reports_and_articles("cholera")
    # sender.send_reports_and_articles("chikungunya")

